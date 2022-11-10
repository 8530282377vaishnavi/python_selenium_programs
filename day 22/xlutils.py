import openpyxl
from openpyxl.styles import PatternFill

def GetRowCount(file,sheetname):
    workbook=openpyxl.workbook(file)
    sheet=workbook[sheetname]
    sheet.max_row

def GetColunmCount(file,sheetname):
    workbook=openpyxl.workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def ReadData(file,sheetname,rownum,columnnum):
    workbook=openpyxl.workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,columnnum).value

def WriteData(file,sheetname,rownum,columnnum,data):
    workbook=openpyxl.workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,columnnum).value=data
    workbook.save()

def FillGreenColour(file,sheetname,rownum, columnnum):
    workbook=openpyxl.workbook(file)
    greenFill=PatternFill(start_color="60b212",
                          end_color="60b212",
                          fill_type="solid")
    sheet=workbook[sheetname]
    sheet.cell(rownum,columnnum).fill=greenFill
    workbook.save()


def FillRedColour(file,sheetname,ronum,columnnum):
    workbook=openpyxl.workbook(file)
    fillRed=PatternFill(start_color="ff0000",
                        end_color="ff0000",
                        fill_type="solid")
    sheet=workbook[sheetname]
    sheet.cell(ronum,columnnum).fill=fillRed
    workbook.save()
