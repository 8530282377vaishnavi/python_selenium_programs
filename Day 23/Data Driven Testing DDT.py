import openpyxl

file="C:\selenium web drivers\Test cases related to Ecommerce web site (1).xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Table 10"]

rows=sheet.max_row # count the no. of rows inexcel sheet
columns=sheet.max_column # count the no. of colums in excel sheet


# Reading the all the rows anfd coluns in excel sheet
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value, end="              ")
    print()
