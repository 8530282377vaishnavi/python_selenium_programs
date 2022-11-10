import openpyxl
# for same data writing
file="C:\\selenium web drivers\\Book1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="welcome"

workbook.save(file)


# different data writing
sheet.cell(1,1).value=123
sheet.cell(1,2).value="vaishnavi"

sheet.cell(2,1).value=124
sheet.cell(2,2).value="Ajinkya"

sheet.cell(3,1).value=125
sheet.cell(3,2).value="vaishnavi"


workbook.save()